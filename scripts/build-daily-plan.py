#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""⛔ 已封存（2026-08-21）——請勿再執行

課程進度的正本已改為 Notion「📅 每日課程進度」，由 scripts/sync-notion.mjs 的
syncDailyPlan() 產生 data/daily-plan.json。老師在 Notion 改完按「立即更新班網」即生效。
本腳本若重跑，會把 Notion 上的最新進度用封存的 xlsx 蓋掉。

保留原因：換學年時可用來把新的進度表 xlsx 轉成一次性匯入的素材
（流程：本腳本 → data/daily-plan.json → scripts/classos/f20-import-daily-plan.mjs 匯入 Notion）。
要這樣用時，請先確認 Notion 那份已無需保留的手改內容。

────────────────────────────────────────────────────────
每日課程進度表（xlsx）→ data/daily-plan.json

用法（在 class-website/ 底下執行）：
    python3 scripts/build-daily-plan.py

輸入：../115學年度四上每日課程進度表.xlsx（正本；老師改這份，改完重跑本腳本）
輸出：data/daily-plan.json（教學駕駛艙行事曆檢視的資料來源）

── 本腳本「不」處理節次對齊 ──────────────────────────────────
「哪一節上哪一科」屬於日課表（data/schedule.json）的事，而日課表是 Notion
「🕐 日課表」每天自動同步下來的，隨時可能變。若在這裡就把進度釘死在「第七節」
這種節次名上，日課表一改、自動同步只更新了一半，駕駛艙就會把數學進度掛到
資訊課那一格去，而且畫面上不會有任何警告（2026-08-07 實測確認）。

所以本腳本只輸出「**每週每科要教的內容，依日期排好**」，
節次對齊改由 assets/js/cockpit.js 在開頁時、依當下的日課表即時計算。
→ **改日課表不必重跑本腳本**；只有改進度表 xlsx 才需要跑。

── 換班級／換學年怎麼沿用 ────────────────────────────────
科目欄不寫死：直接讀表頭第 4 欄到「重要行事與備註」之前的所有欄，
欄名取「課程進度」之前的字（「國語課程進度（5節/週）」→「國語」）。
換成別的進度表，只要保持「週次｜日期｜星期｜<各科>｜備註」這個欄位結構就能直接用。
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent          # class-website/
XLSX = ROOT.parent / "115學年度四上每日課程進度表.xlsx"  # 班級事務/
OUT = ROOT / "data" / "daily-plan.json"

SHEET = "每日課程進度表"
HOLIDAY_MARK = "【放假不排新課】"
DOW_NAME = {"星期一": 1, "星期二": 2, "星期三": 3, "星期四": 4, "星期五": 5}
NOTE_HEADER = "重要行事與備註"


def blank(text: str) -> bool:
    """空白／破折號／放假 → 不是實質教學內容"""
    return (text or "").strip() in ("", "-", "—", "－", HOLIDAY_MARK)


def find_header(ws):
    """找表頭列：前幾列是標題／規則說明／空白間隔，真正的表頭是含「週次…星期」那一列。
       自動偵測而不寫死列號，換一份排版稍有不同的進度表也能直接用。"""
    for row in ws.iter_rows(min_row=1, max_row=15):
        cells = [str(c.value or "").strip() for c in row]
        if len(cells) >= 3 and "週次" in cells[0] and "星期" in cells[2]:
            return row[0].row, cells
    sys.exit("❌ 找不到表頭列（需要一列的前三欄是「週次｜日期｜星期」）")


def read_columns(header):
    """科目欄＝第 4 欄起、到「重要行事與備註」為止；欄名取「課程進度」之前的字"""
    cols = []
    for i, name in enumerate(header):
        if i < 3 or not name or NOTE_HEADER in name:
            continue
        cols.append((i, re.split(r"課程進度|進度", name)[0].strip() or name))
    if not cols:
        sys.exit("❌ 表頭讀不到任何科目欄，請確認第 4 欄之後的欄位名稱")
    return cols


def main():
    if not XLSX.exists():
        sys.exit(f"❌ 找不到進度表正本：{XLSX}")
    ws = openpyxl.load_workbook(XLSX, data_only=True)[SHEET]
    header_row, header = find_header(ws)
    subject_cols = read_columns(header)
    ni = next((i for i, n in enumerate(header) if NOTE_HEADER in n), None)

    days, weeks = [], {}
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(r)
        week_s, date_v, dow_s = (row + [None] * 3)[:3]
        if not week_s or not date_v:
            continue
        m = re.search(r"(\d+)", str(week_s))
        if not m:
            continue
        week = int(m.group(1))
        date = date_v.date() if hasattr(date_v, "date") else datetime.strptime(str(date_v), "%Y/%m/%d").date()
        iso = date.isoformat()

        texts = {name: (str(row[i]).strip() if i < len(row) and row[i] else "")
                 for i, name in subject_cols}
        note = (str(row[ni]).strip() if ni is not None and ni < len(row) and row[ni] else "")

        days.append({
            "date": iso,
            "week": week,
            "dow": DOW_NAME.get(str(dow_s).strip(), date.isoweekday()),
            # 放假日＝所有科目都寫【放假不排新課】；當天不佔節次，也不吃掉該週的進度條目
            "holiday": all(t == HOLIDAY_MARK for t in texts.values()) and bool(texts),
            "note": note,
        })

        wk = weeks.setdefault(str(week), {})
        for name, text in texts.items():
            if not blank(text):
                wk.setdefault(name, []).append({"date": iso, "text": text})

    days.sort(key=lambda d: d["date"])
    for wk in weeks.values():
        for entries in wk.values():
            entries.sort(key=lambda e: e["date"])

    payload = {
        "meta": {
            "source": XLSX.name,
            "builtAt": datetime.now().isoformat(timespec="seconds"),
            "subjects": [name for _, name in subject_cols],
            "days": len(days),
            "weeks": len(weeks),
            "note": "由 scripts/build-daily-plan.py 產生，請勿手改。"
                    "本檔不含節次對齊——節次由 cockpit.js 依當下的 data/schedule.json 即時計算，"
                    "所以改日課表不必重跑本腳本，只有改進度表 xlsx 才要跑。",
        },
        # 每個上課日的日期／週次／星期／是否放假／行事備註
        "days": days,
        # 每週每科要教的內容（依日期排好，未對齊節次）
        "weeks": weeks,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    total = sum(len(e) for wk in weeks.values() for e in wk.values())
    print(f"✅ {OUT.relative_to(ROOT)}：{len(days)} 個上課日／{len(weeks)} 週，"
          f"科目 {'、'.join(n for _, n in subject_cols)}，共 {total} 條進度，"
          f"放假日 {sum(1 for d in days if d['holiday'])} 天")


if __name__ == "__main__":
    main()
